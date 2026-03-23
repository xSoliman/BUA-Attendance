"""
Core attendance sheet generation logic.
Supports multiple courses — each becomes a separate tab in one workbook.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _write_course_sheet(ws, students: list[dict]) -> None:
    """Write a styled attendance sheet into an existing openpyxl worksheet."""
    students = sorted(students, key=lambda s: (s.get("section", ""), str(s.get("id", ""))))

    headers = [
        "ID", "Name", "Section",
        "Week 1", "Week 2", "Week 3", "Week 4", "Week 5",
        "Week 6", "Week 7", "Week 8", "Week 9", "Week 10",
        "Total Attendance",
    ]

    header_fill  = PatternFill(start_color="0F45A8", end_color="0F45A8", fill_type="solid")
    header_font  = Font(bold=True, color="FFFFFF")
    center       = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )
    color_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    color_gray  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value      = header
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = center

    # Student rows
    for idx, student in enumerate(students, start=2):
        row_fill = color_white if idx % 2 == 0 else color_gray

        def styled(col, value, bold=False):
            c = ws.cell(row=idx, column=col)
            c.value     = value
            c.alignment = center
            c.fill      = row_fill
            c.border    = thin_border
            if bold:
                c.font = Font(bold=True)

        styled(1, str(student.get("id", "")))
        styled(2, str(student.get("name", "")))
        styled(3, str(student.get("section", "")))
        for col in range(4, 14):
            styled(col, None)
        styled(14, (
            f"=SUMPRODUCT((UPPER(D{idx}:M{idx})=\"P\")"
            f"+(D{idx}:M{idx}=1))"
        ), bold=True)

    # 10 extra blank rows
    last_row = len(students) + 1
    for extra in range(1, 11):
        row = last_row + extra
        row_fill = color_white if row % 2 == 0 else color_gray
        for col in range(1, 15):
            c = ws.cell(row=row, column=col)
            c.alignment = center
            c.fill      = row_fill
            c.border    = thin_border
            if col == 14:
                c.value = (
                    f"=SUMPRODUCT((UPPER(D{row}:M{row})=\"P\")"
                    f"+(D{row}:M{row}=1))"
                )
                c.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 13.0
    ws.column_dimensions["B"].width = 31.25
    ws.column_dimensions["C"].width = 13.0
    for col in range(4, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13.0
    ws.column_dimensions["N"].width = 18.0

    ws.freeze_panes = "D2"


def generate_attendance_sheet(courses: list[dict]) -> bytes:
    """
    Generate a multi-tab attendance workbook.

    Each course dict: { "name": str, "students": list[dict] }
    Each student dict: { "id": str, "name": str, "section": str (optional) }

    Returns raw .xlsx bytes.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for course in courses:
        tab_name = (course.get("name") or "Course").strip()
        # Excel sheet names: max 31 chars, no special chars
        tab_name = tab_name[:31]
        for ch in r"\/*?:[]":
            tab_name = tab_name.replace(ch, "")
        tab_name = tab_name or "Course"

        ws = wb.create_sheet(title=tab_name)
        _write_course_sheet(ws, course.get("students", []))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
