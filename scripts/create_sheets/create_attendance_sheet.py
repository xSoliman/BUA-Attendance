#!/usr/bin/env python3
"""
Attendance Sheet Generator

Creates attendance sheets for specific sections by filtering students from input Excel file.

Usage:
    python create_attendance_sheet.py <sheet_name> <section1> <section2> ...

Examples:
    python create_attendance_sheet.py sampleInput A1
    python create_attendance_sheet.py AI-Students A1 A2 B7
    python create_attendance_sheet.py AS-Students A1 A2 A3 A4 A5

Input:
    - Reads from: input/<sheet_name>.xlsx
    - Expected columns: ID, Name, Section

Output:
    - Creates: output/attendance_<sections>.xlsx
    - Format matches: sampleOutput.xlsx
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


def get_script_directory():
    """Get the directory where the script is located."""
    return os.path.dirname(os.path.abspath(__file__))


def load_students(input_file, sections):
    """
    Load students from Excel file filtered by sections.
    
    Args:
        input_file: Path to input Excel file
        sections: List of section codes to filter (e.g., ['A1', 'A2', 'B7'])
        
    Returns:
        DataFrame with filtered students
    """
    if not os.path.exists(input_file):
        print(f"Error: Input file not found: {input_file}")
        sys.exit(1)
    
    # Read all sheets from the Excel file
    excel_file = pd.ExcelFile(input_file)
    all_students = []
    
    print(f"Reading from: {os.path.basename(input_file)}")
    print(f"Looking for sections: {', '.join(sections)}")
    
    for sheet_name in excel_file.sheet_names:
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        
        # Check if required columns exist
        if 'Section' not in df.columns:
            continue
        
        # Filter by sections
        filtered = df[df['Section'].isin(sections)]
        
        if len(filtered) > 0:
            print(f"  Found {len(filtered)} students in sheet '{sheet_name}'")
            all_students.append(filtered)
    
    if not all_students:
        print(f"\nError: No students found for sections: {', '.join(sections)}")
        print("Please check:")
        print("  1. Section codes are correct (case-sensitive)")
        print("  2. Input file has a 'Section' column")
        print("  3. Students exist for these sections")
        sys.exit(1)
    
    # Combine all filtered students
    combined_df = pd.concat(all_students, ignore_index=True)
    
    # Sort by Section, then by ID
    if 'ID' in combined_df.columns:
        combined_df = combined_df.sort_values(['Section', 'ID'])
    elif 'Student_ID' in combined_df.columns:
        combined_df = combined_df.sort_values(['Section', 'Student_ID'])
    
    print(f"\nTotal students found: {len(combined_df)}")
    
    return combined_df


def create_attendance_sheet(students_df, output_file, sections):
    """
    Create attendance sheet matching sampleOutput.xlsx format with Total Attendance column.
    
    Args:
        students_df: DataFrame with student data
        output_file: Path to output file
        sections: List of sections being processed
    """
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Create headers (including Total Attendance)
    headers = ['ID', 'Name', 'Section', 'Week 1', 'Week 2', 'Week 3', 'Week 4', 
               'Week 5', 'Week 6', 'Week 7', 'Week 8', 'Week 9', 'Week 10', 'Total Attendance']
    
    # Header styling (matching sampleOutput.xlsx)
    header_fill = PatternFill(start_color="0F45A8", end_color="0F45A8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers with formatting
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Get ID column name
    id_column = 'ID' if 'ID' in students_df.columns else 'Student_ID'
    
    # Write student data
    for idx, (_, student) in enumerate(students_df.iterrows(), start=2):
        row = idx
        
        # ID
        ws.cell(row=row, column=1).value = int(student[id_column])
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="bottom")
        
        # Name
        if 'Name' in students_df.columns:
            ws.cell(row=row, column=2).value = str(student['Name'])
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="bottom")
        
        # Section
        ws.cell(row=row, column=3).value = str(student['Section'])
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="bottom")
        
        # Week columns (empty)
        for col in range(4, 14):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="bottom")
        
        # Total Attendance column (column 14) - Formula to count P, p, or 1
        # Formula: SUMPRODUCT((UPPER(D2:M2)="P")+(D2:M2="1"))
        week_start_col = 'D'  # Week 1 starts at column D
        week_end_col = 'M'    # Week 10 ends at column M
        formula = f'=SUMPRODUCT((UPPER({week_start_col}{row}:{week_end_col}{row})="P")+({week_start_col}{row}:{week_end_col}{row}=1))'
        
        total_cell = ws.cell(row=row, column=14)
        total_cell.value = formula
        total_cell.alignment = Alignment(horizontal="center", vertical="bottom")
        total_cell.font = Font(bold=True)
    
    # Set column widths (matching sampleOutput.xlsx)
    ws.column_dimensions['A'].width = 13.0   # ID
    ws.column_dimensions['B'].width = 31.25  # Name
    ws.column_dimensions['C'].width = 13.0   # Section
    
    # Week columns
    for col in range(4, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13.0
    
    # Total Attendance column
    ws.column_dimensions['N'].width = 18.0
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Created attendance sheet: {output_file}")
    print(f"  Sections: {', '.join(sections)}")
    print(f"  Students: {len(students_df)}")
    print(f"  Total Attendance column added with auto-calculation")


def main():
    """Main entry point for the attendance sheet generator."""
    if len(sys.argv) < 3:
        print("Attendance Sheet Generator")
        print("=" * 60)
        print("\nUsage: python create_attendance_sheet.py <sheet_name> <section1> <section2> ...")
        print("\nExamples:")
        print("  python create_attendance_sheet.py sampleInput A1")
        print("  python create_attendance_sheet.py AI-Students A1 A2 B7")
        print("  python create_attendance_sheet.py AS-Students A1 A2 A3 A4 A5")
        print("\nInput:")
        print("  Reads from: input/<sheet_name>.xlsx")
        print("\nOutput:")
        print("  Creates: output/attendance_<sections>.xlsx")
        print("  Example: output/attendance_A1_A2_B7.xlsx")
        sys.exit(1)
    
    # Get sheet name and sections from command line arguments
    sheet_name = sys.argv[1]
    sections = [arg.upper() for arg in sys.argv[2:]]
    
    # Setup paths
    script_dir = get_script_directory()
    input_dir = os.path.join(script_dir, "input")
    input_file = os.path.join(input_dir, f"{sheet_name}.xlsx")
    output_dir = os.path.join(script_dir, "output")
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Create output filename
    sections_str = "_".join(sections)
    output_file = os.path.join(output_dir, f"attendance_{sections_str}.xlsx")
    
    print("=" * 60)
    print("Attendance Sheet Generator")
    print("=" * 60)
    
    # Load students
    students_df = load_students(input_file, sections)
    
    # Create attendance sheet
    create_attendance_sheet(students_df, output_file, sections)
    
    print("=" * 60)
    print("Done!")
    print("=" * 60)


if __name__ == "__main__":
    main()
