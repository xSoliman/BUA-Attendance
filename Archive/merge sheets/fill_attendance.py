#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Attendance Marking Script
Reads source.xlsx and fills target.xlsx with 'P' marks where source has '1'
Handles Arabic name variations and partial matches
"""

import pandas as pd
import openpyxl
import re
from difflib import SequenceMatcher
import unicodedata


def normalize_arabic_name(name):
    """
    Normalize Arabic names by:
    - Converting to lowercase
    - Removing extra spaces
    - Normalizing Arabic characters (أ, إ, آ -> ا), (ى -> ي), (ة -> ه)
    - Removing diacritics
    """
    if pd.isna(name):
        return ""
    
    name = str(name).strip()
    
    # Remove Arabic diacritics (tashkeel)
    name = re.sub(r'[\u064B-\u065F\u0670]', '', name)
    
    # Normalize Arabic characters
    replacements = {
        'أ': 'ا', 'إ': 'ا', 'آ': 'ا', 'ٱ': 'ا',
        'ى': 'ي', 'ئ': 'ي',
        'ة': 'ه',
        'ؤ': 'و'
    }
    
    for old, new in replacements.items():
        name = name.replace(old, new)
    
    # Remove extra whitespace
    name = ' '.join(name.split())
    
    return name


def get_name_tokens(name):
    """Split name into tokens (words)"""
    normalized = normalize_arabic_name(name)
    return normalized.split()


def calculate_match_score(source_name, target_name):
    """
    Calculate match score between two names
    Returns: (score, match_type)
    - score: 0-100
    - match_type: 'exact', 'partial', 'fuzzy', or 'no_match'
    """
    source_norm = normalize_arabic_name(source_name)
    target_norm = normalize_arabic_name(target_name)
    
    # Exact match
    if source_norm == target_norm:
        return 100, 'exact'
    
    source_tokens = get_name_tokens(source_name)
    target_tokens = get_name_tokens(target_name)
    
    # Check if one name is a subset of the other (partial match)
    # e.g., "احمد محمد علي" matches "احمد محمد علي صبري"
    if len(source_tokens) >= 2 and len(target_tokens) >= 2:
        # Check if shorter name is contained in longer name
        shorter = source_tokens if len(source_tokens) <= len(target_tokens) else target_tokens
        longer = target_tokens if len(source_tokens) <= len(target_tokens) else source_tokens
        
        # Check if all tokens of shorter name appear in order in longer name
        matches = 0
        longer_idx = 0
        for token in shorter:
            for i in range(longer_idx, len(longer)):
                if token == longer[i]:
                    matches += 1
                    longer_idx = i + 1
                    break
        
        if matches == len(shorter):
            # All tokens matched in order - this is a strong partial match
            # Give higher score if most tokens match
            ratio = matches / max(len(source_tokens), len(target_tokens))
            # Boost score for partial matches (minimum 85 if all shorter tokens match)
            score = max(85, int(ratio * 95))
            return score, 'partial'
    
    # Fuzzy matching using sequence matcher
    similarity = SequenceMatcher(None, source_norm, target_norm).ratio()
    score = int(similarity * 100)
    
    if score >= 85:
        return score, 'fuzzy'
    
    return score, 'no_match'


def find_best_match(source_name, target_names_list, threshold=75):
    """
    Find the best matching name from target list
    Returns: (index, score, match_type) or (None, 0, 'no_match')
    """
    best_match = None
    best_score = 0
    best_type = 'no_match'
    matches_found = []
    
    for idx, target_name in enumerate(target_names_list):
        score, match_type = calculate_match_score(source_name, target_name)
        
        if score >= threshold:
            matches_found.append((idx, score, match_type))
            if score > best_score:
                best_score = score
                best_match = idx
                best_type = match_type
    
    # Only return match if there's exactly one good match or one clear best match
    if len(matches_found) == 1:
        return matches_found[0]
    elif len(matches_found) > 1:
        # Check if best match is significantly better than others
        sorted_matches = sorted(matches_found, key=lambda x: x[1], reverse=True)
        if sorted_matches[0][1] - sorted_matches[1][1] >= 5:  # 5 point difference
            return sorted_matches[0]
        else:
            # Multiple similar matches - ambiguous
            return None, 0, 'ambiguous'
    
    return None, 0, 'no_match'


def main():
    print("Loading Excel files...")
    
    # Load all sheets from source
    source_xl = pd.ExcelFile('source.xlsx')
    source_sheets = source_xl.sheet_names
    
    # Open target workbook with openpyxl for in-place editing
    target_wb = openpyxl.load_workbook('target.xlsx')
    target_sheets = target_wb.sheetnames
    
    print(f"Source sheets: {source_sheets}")
    print(f"Target sheets: {target_sheets}")
    print()
    
    # Overall statistics
    overall_stats = {
        'total_source_entries': 0,
        'matched': 0,
        'not_found': [],
        'ambiguous': [],
        'marked_cells': 0,
        'sheets_processed': 0
    }
    
    # Process each source sheet
    for source_sheet_name in source_sheets:
        print("\n" + "="*60)
        print(f"Processing source sheet: {source_sheet_name}")
        print("="*60)
        
        source_df = pd.read_excel('source.xlsx', sheet_name=source_sheet_name)
        print(f"Source entries in this sheet: {len(source_df)}")
        
        # Find matching target sheet (try exact match first, then fuzzy)
        target_sheet_name = None
        target_ws = None
        
        # Try exact match
        if source_sheet_name in target_sheets:
            target_sheet_name = source_sheet_name
            target_ws = target_wb[source_sheet_name]
        else:
            # Try fuzzy match (normalize names)
            source_normalized = source_sheet_name.lower().replace(' ', '').replace('&', '')
            for tgt_sheet in target_sheets:
                tgt_normalized = tgt_sheet.lower().replace(' ', '').replace('&', '')
                if source_normalized == tgt_normalized:
                    target_sheet_name = tgt_sheet
                    target_ws = target_wb[tgt_sheet]
                    break
        
        if target_ws is None:
            print(f"⚠ Warning: No matching target sheet found for '{source_sheet_name}'")
            print(f"   Skipping this sheet...")
            continue
        
        # Load ONLY this specific target sheet for matching
        target_df = pd.read_excel('target.xlsx', sheet_name=target_sheet_name)
        print(f"Matched to target sheet: '{target_sheet_name}' ({len(target_df)} rows)")
        
        # Get week columns (excluding 'Name' column)
        week_columns = [col for col in source_df.columns if col != 'Name']
        
        # Build column index mapping for target sheet
        target_header_row = 1  # Excel is 1-indexed
        target_col_map = {}
        for col_idx, cell in enumerate(target_ws[target_header_row], start=1):
            if cell.value in week_columns:
                target_col_map[cell.value] = col_idx
        
        # Sheet statistics
        sheet_stats = {
            'matched': 0,
            'not_found': [],
            'ambiguous': [],
            'marked_cells': 0
        }
        
        # Process each person in source
        for src_idx, src_row in source_df.iterrows():
            source_name = src_row['Name']
            
            if pd.isna(source_name):
                continue
            
            overall_stats['total_source_entries'] += 1
            
            # Find matching person ONLY in this specific target sheet
            target_idx, score, match_type = find_best_match(
                source_name, 
                target_df['Name'].tolist(),
                threshold=75
            )
            
            if target_idx is not None and match_type != 'ambiguous':
                target_name = target_df.loc[target_idx, 'Name']
                sheet_stats['matched'] += 1
                
                # Calculate Excel row (pandas index + 2, because Excel is 1-indexed and has header)
                excel_row = target_idx + 2
                
                # Mark attendance for each week
                for week_col in week_columns:
                    if week_col in target_col_map:
                        source_value = src_row[week_col]
                        if source_value == 1 or source_value == 1.0:
                            # Write 'P' directly to Excel cell
                            excel_col = target_col_map[week_col]
                            target_ws.cell(row=excel_row, column=excel_col, value='P')
                            sheet_stats['marked_cells'] += 1
                
                print(f"✓ Matched: '{source_name}' -> '{target_name}' (score: {score}, type: {match_type})")
            
            elif match_type == 'ambiguous':
                sheet_stats['ambiguous'].append((source_sheet_name, source_name))
                print(f"⚠ Ambiguous: '{source_name}' (multiple similar matches found)")
            
            else:
                sheet_stats['not_found'].append((source_sheet_name, source_name))
                print(f"✗ Not found: '{source_name}'")
        
        # Update overall stats
        overall_stats['matched'] += sheet_stats['matched']
        overall_stats['not_found'].extend(sheet_stats['not_found'])
        overall_stats['ambiguous'].extend(sheet_stats['ambiguous'])
        overall_stats['marked_cells'] += sheet_stats['marked_cells']
        overall_stats['sheets_processed'] += 1
        
        # Print sheet summary
        print(f"\nSheet '{source_sheet_name}' summary:")
        print(f"  Matched: {sheet_stats['matched']}")
        print(f"  Not found: {len(sheet_stats['not_found'])}")
        print(f"  Ambiguous: {len(sheet_stats['ambiguous'])}")
        print(f"  Cells marked: {sheet_stats['marked_cells']}")
    
    # Save the workbook (modifies the original file)
    print("\n" + "="*60)
    print("Saving changes to target.xlsx...")
    print("="*60)
    target_wb.save('target.xlsx')
    print("✓ File updated successfully")
    
    # Generate overall report
    print("\n" + "="*60)
    print("OVERALL ATTENDANCE MARKING REPORT")
    print("="*60)
    print(f"Sheets processed: {overall_stats['sheets_processed']}")
    print(f"Total source entries: {overall_stats['total_source_entries']}")
    print(f"Successfully matched: {overall_stats['matched']}")
    print(f"Total cells marked with 'P': {overall_stats['marked_cells']}")
    print(f"Not found: {len(overall_stats['not_found'])}")
    print(f"Ambiguous matches: {len(overall_stats['ambiguous'])}")
    
    if overall_stats['not_found']:
        print("\n--- Names NOT FOUND in target (by sheet) ---")
        current_sheet = None
        for sheet_name, name in overall_stats['not_found']:
            if sheet_name != current_sheet:
                print(f"\n[{sheet_name}]")
                current_sheet = sheet_name
            print(f"  • {name}")
    
    if overall_stats['ambiguous']:
        print("\n--- AMBIGUOUS matches (by sheet) ---")
        current_sheet = None
        for sheet_name, name in overall_stats['ambiguous']:
            if sheet_name != current_sheet:
                print(f"\n[{sheet_name}]")
                current_sheet = sheet_name
            print(f"  • {name}")
    
    print("\n✓ Process completed successfully!")
    print(f"✓ Original target.xlsx file has been updated in place")
    print(f"✓ Each source sheet was matched ONLY with its corresponding target sheet")
    
    # Generate detailed report file
    print("\n" + "="*60)
    print("Generating detailed report file...")
    print("="*60)
    
    with open('attendance_report.txt', 'w', encoding='utf-8') as f:
        f.write("="*60 + "\n")
        f.write("ATTENDANCE MARKING DETAILED REPORT\n")
        f.write("="*60 + "\n\n")
        
        f.write(f"Sheets processed: {overall_stats['sheets_processed']}\n")
        f.write(f"Total source entries: {overall_stats['total_source_entries']}\n")
        f.write(f"Successfully matched: {overall_stats['matched']}\n")
        f.write(f"Total cells marked with 'P': {overall_stats['marked_cells']}\n")
        f.write(f"Not found: {len(overall_stats['not_found'])}\n")
        f.write(f"Ambiguous matches: {len(overall_stats['ambiguous'])}\n\n")
        
        # Group not found and ambiguous by sheet
        not_found_by_sheet = {}
        ambiguous_by_sheet = {}
        
        for sheet_name, name in overall_stats['not_found']:
            if sheet_name not in not_found_by_sheet:
                not_found_by_sheet[sheet_name] = []
            not_found_by_sheet[sheet_name].append(name)
        
        for sheet_name, name in overall_stats['ambiguous']:
            if sheet_name not in ambiguous_by_sheet:
                ambiguous_by_sheet[sheet_name] = []
            ambiguous_by_sheet[sheet_name].append(name)
        
        # Write detailed report for each sheet
        f.write("="*60 + "\n")
        f.write("DETAILED REPORT BY SHEET\n")
        f.write("="*60 + "\n\n")
        
        for sheet_name in source_sheets:
            f.write("-" * 60 + "\n")
            f.write(f"Sheet: {sheet_name}\n")
            f.write("-" * 60 + "\n\n")
            
            # Not found names for this sheet
            if sheet_name in not_found_by_sheet:
                f.write(f"NOT FOUND ({len(not_found_by_sheet[sheet_name])} names):\n")
                for name in not_found_by_sheet[sheet_name]:
                    f.write(f"  • {name}\n")
                f.write("\n")
            else:
                f.write("NOT FOUND: None\n\n")
            
            # Ambiguous names for this sheet
            if sheet_name in ambiguous_by_sheet:
                f.write(f"AMBIGUOUS MATCHES ({len(ambiguous_by_sheet[sheet_name])} names):\n")
                for name in ambiguous_by_sheet[sheet_name]:
                    f.write(f"  • {name}\n")
                f.write("\n")
            else:
                f.write("AMBIGUOUS MATCHES: None\n\n")
        
        f.write("="*60 + "\n")
        f.write("END OF REPORT\n")
        f.write("="*60 + "\n")
    
    print("✓ Detailed report saved to: attendance_report.txt")


if __name__ == "__main__":
    main()
